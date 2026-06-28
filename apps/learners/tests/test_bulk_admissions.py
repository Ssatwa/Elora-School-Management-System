from datetime import date
from io import BytesIO
from typing import cast

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook

from apps.academics.models import AcademicYear, Grade, Stream
from apps.accounts.models import Membership
from apps.learners.models import Enrollment, Learner
from apps.learners.services.bulk_admissions import (
    build_csv_template,
    bulk_admission_headers,
    parse_bulk_admission_upload,
)
from apps.tenancy.models import School
from tests.factories import MembershipFactory, SchoolFactory

pytestmark = pytest.mark.django_db


def login_membership(client, school: School, role_code: str):
    membership = cast(
        Membership,
        MembershipFactory(school=school, role_code=role_code),
    )
    client.force_login(membership.user)
    return membership


def create_academic_context(school: School):
    year = AcademicYear.objects.create(
        school=school,
        name="2026",
        start_date=date(2026, 1, 1),
        end_date=date(2026, 12, 31),
        status=AcademicYear.Status.ACTIVE,
    )
    grade = Grade.objects.create(
        school=school,
        code="G7",
        name="Grade 7",
        education_level=Grade.EducationLevel.JUNIOR_SCHOOL,
        order=7,
    )
    stream = Stream.objects.create(
        school=school,
        grade=grade,
        code="E",
        name="East",
    )
    return year, grade, stream


def csv_upload(content: str, name: str = "learners.csv"):
    return SimpleUploadedFile(
        name,
        content.encode(),
        content_type="text/csv",
    )


def valid_csv(*, year: AcademicYear, grade: Grade, stream: Stream):
    return "\n".join(
        [
            ",".join(bulk_admission_headers()),
            ",".join(
                [
                    "Amina",
                    "",
                    "Kamau",
                    "2013-06-12",
                    "female",
                    year.name,
                    grade.name,
                    stream.name,
                    "2026-01-06",
                    "Wanjiku",
                    "Kamau",
                    "parent@example.test",
                    "+254700000001",
                    "mother",
                    "O+",
                    "Peanuts",
                    "",
                    "",
                ]
            ),
        ]
    )


def test_bulk_template_headers_match_admission_form_labels():
    assert bulk_admission_headers() == [
        "First name",
        "Middle name",
        "Last name",
        "Date of birth",
        "Gender",
        "Academic year",
        "Grade",
        "Stream",
        "Admission date",
        "Guardian first name",
        "Guardian last name",
        "Guardian email",
        "Guardian phone number",
        "Guardian relationship",
        "Blood group",
        "Allergies",
        "Conditions",
        "Medication",
    ]
    assert build_csv_template().splitlines()[0] == ",".join(bulk_admission_headers())


def test_parse_bulk_admission_upload_validates_with_admission_form():
    school = cast(School, SchoolFactory())
    year, grade, stream = create_academic_context(school)

    result = parse_bulk_admission_upload(
        uploaded_file=csv_upload(valid_csv(year=year, grade=grade, stream=stream)),
        school=school,
    )

    assert result.has_errors is False
    assert result.valid_count == 1
    assert result.rows[0].cleaned_data["grade"] == grade
    assert result.rows[0].cleaned_data["stream"] == stream


def test_parse_bulk_admission_upload_reports_missing_invalid_and_duplicate_rows():
    school = cast(School, SchoolFactory())
    year, grade, stream = create_academic_context(school)
    content = "\n".join(
        [
            ",".join(bulk_admission_headers()),
            f"Amina,,Kamau,2013-06-12,female,{year.name},{grade.name},{stream.name},2026-01-06,Wanjiku,Kamau,parent@example.test,+254700000001,mother,O+,,,",
            f"Amina,,Kamau,2013-06-12,female,{year.name},{grade.name},{stream.name},not-a-date,Wanjiku,Kamau,parent@example.test,+254700000001,mother,O+,,,",
        ]
    )

    result = parse_bulk_admission_upload(
        uploaded_file=csv_upload(content),
        school=school,
    )

    assert result.has_errors is True
    assert result.valid_count == 1
    assert result.error_count == 1
    assert result.rows[1].row_number == 3
    assert any(error.field == "Admission date" for error in result.rows[1].errors)
    assert any("Duplicate learner" in error.issue for error in result.rows[1].errors)


def test_parse_bulk_admission_upload_rejects_unknown_grade_for_active_school():
    school = cast(School, SchoolFactory())
    other_school = cast(School, SchoolFactory())
    year, _, _ = create_academic_context(school)
    _, other_grade, other_stream = create_academic_context(other_school)
    other_grade.name = "Secret Grade"
    other_grade.code = "SECRET-G"
    other_grade.save(update_fields=["name", "code"])
    other_stream.name = "Secret Stream"
    other_stream.code = "SECRET-S"
    other_stream.save(update_fields=["name", "code"])
    content = valid_csv(year=year, grade=other_grade, stream=other_stream)

    result = parse_bulk_admission_upload(
        uploaded_file=csv_upload(content),
        school=school,
    )

    assert result.has_errors is True
    assert any(error.field == "Grade" for error in result.rows[0].errors)
    assert any(error.field == "Stream" for error in result.rows[0].errors)


def test_bulk_upload_preview_and_confirm_create_only_valid_learners(client):
    school = cast(School, SchoolFactory())
    login_membership(client, school, "school_admin")
    year, grade, stream = create_academic_context(school)
    content = "\n".join(
        [
            ",".join(bulk_admission_headers()),
            f"Amina,,Kamau,2013-06-12,female,{year.name},{grade.name},{stream.name},2026-01-06,Wanjiku,Kamau,parent@example.test,+254700000001,mother,O+,,,",
            f"Brian,,Otieno,2013-06-12,wrong,{year.name},{grade.name},{stream.name},2026-01-06,Wanjiku,Otieno,parent2@example.test,+254700000002,mother,O+,,,",
        ]
    )

    preview = client.post(
        reverse("learners:bulk_admit"),
        {"file": csv_upload(content)},
        HTTP_HOST=f"{school.slug}.localhost",
    )

    assert preview.status_code == 200
    preview_content = preview.content.decode()
    assert "Amina Kamau" in preview_content
    assert "Select a valid choice" in preview_content
    assert "data-bulk-confirm" in preview_content

    confirm = client.post(
        reverse("learners:bulk_admit_confirm"),
        HTTP_HOST=f"{school.slug}.localhost",
    )

    assert confirm.status_code == 302
    assert Learner.objects.for_school(school).count() == 1
    learner = Learner.objects.for_school(school).get()
    assert learner.full_name == "Amina Kamau"
    assert Enrollment.objects.get(learner=learner).stream == stream


def test_bulk_upload_page_marks_template_link_as_download(client):
    school = cast(School, SchoolFactory())
    login_membership(client, school, "school_admin")

    response = client.get(
        reverse("learners:bulk_admit"),
        HTTP_HOST=f"{school.slug}.localhost",
    )

    content = response.content.decode()
    assert response.status_code == 200
    assert 'href="/learners/admit/bulk/template/"' in content
    assert 'download="learner-admission-template.xlsx"' in content


def test_bulk_upload_template_downloads_excel_workbook(client):
    school = cast(School, SchoolFactory())
    login_membership(client, school, "principal")

    response = client.get(
        reverse("learners:bulk_admit_template"),
        HTTP_HOST=f"{school.slug}.localhost",
    )

    assert response.status_code == 200
    assert response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"] == (
        'attachment; filename="learner-admission-template.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    sample = [cell.value for cell in sheet[2]]

    assert headers == bulk_admission_headers()
    assert sample[0] == "Sample"


def test_parse_bulk_admission_upload_accepts_xlsx_file():
    from openpyxl import Workbook

    school = cast(School, SchoolFactory())
    year, grade, stream = create_academic_context(school)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(bulk_admission_headers())
    sheet.append(
        [
            "Amina",
            "",
            "Kamau",
            "2013-06-12",
            "female",
            year.name,
            grade.name,
            stream.name,
            "2026-01-06",
            "Wanjiku",
            "Kamau",
            "parent@example.test",
            "+254700000001",
            "mother",
            "O+",
            "Peanuts",
            "",
            "",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    result = parse_bulk_admission_upload(
        uploaded_file=SimpleUploadedFile(
            "learners.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        school=school,
    )

    assert result.has_errors is False
    assert result.valid_count == 1
