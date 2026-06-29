import csv
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path
from uuid import UUID

from django.core.exceptions import ValidationError
from django.db import transaction
from django.forms.utils import pretty_name

from apps.learners.forms import AdmissionForm
from apps.learners.models import AdmissionApplication, Learner
from apps.learners.services.admissions import admit_learner

MAX_UPLOAD_SIZE = 2 * 1024 * 1024
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
SESSION_KEY = "learners_bulk_admission_valid_rows"


@dataclass
class BulkAdmissionError:
    field: str
    issue: str


@dataclass
class BulkAdmissionRow:
    row_number: int
    values: dict[str, str]
    cleaned_data: dict = field(default_factory=dict)
    errors: list[BulkAdmissionError] = field(default_factory=list)

    @property
    def is_valid(self):
        return not self.errors

    @property
    def learner_name(self):
        parts = (
            self.values.get("First name", ""),
            self.values.get("Middle name", ""),
            self.values.get("Last name", ""),
        )
        return " ".join(part for part in parts if part).strip()


@dataclass
class BulkAdmissionResult:
    rows: list[BulkAdmissionRow]
    file_errors: list[str] = field(default_factory=list)

    @property
    def has_errors(self):
        return bool(self.file_errors) or any(not row.is_valid for row in self.rows)

    @property
    def valid_rows(self):
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self):
        return [row for row in self.rows if not row.is_valid]

    @property
    def valid_count(self):
        return len(self.valid_rows)

    @property
    def error_count(self):
        return len(self.error_rows)


def _admission_form():
    return AdmissionForm(school=None)


def bulk_admission_fields():
    return _admission_form().fields


def bulk_admission_headers():
    return [
        field.label if field.label is not None else pretty_name(name)
        for name, field in bulk_admission_fields().items()
    ]


def build_csv_template():
    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(bulk_admission_headers())
    return output.getvalue()


def build_xlsx_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Learners"
    headers = bulk_admission_headers()
    sheet.append(headers)
    sheet.append(
        [
            "Sample",
            "",
            "Learner",
            "2013-06-12",
            "female",
            "2026",
            "Grade 7",
            "East",
            "2026-01-06",
            "Sample",
            "Guardian",
            "guardian@example.test",
            "+254700000001",
            "mother",
            "O+",
            "",
            "",
            "",
        ]
    )
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for column_index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.fill = header_fill
        sheet.column_dimensions[cell.column_letter].width = max(
            14,
            len(headers[column_index - 1]) + 2,
        )
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def serialize_cleaned_row(cleaned_data):
    serialized = {}
    for key, value in cleaned_data.items():
        if hasattr(value, "id"):
            serialized[key] = str(value.id)
        elif hasattr(value, "isoformat"):
            serialized[key] = value.isoformat()
        else:
            serialized[key] = value
    return serialized


def deserialize_session_rows(*, school, rows):
    cleaned_rows = []
    for row in rows:
        form = AdmissionForm(data=row, school=school)
        if not form.is_valid():
            raise ValidationError("The saved upload preview is no longer valid.")
        cleaned_rows.append(form.cleaned_data)
    return cleaned_rows


def _read_csv(uploaded_file):
    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            decoded = ""
    if not decoded:
        raise ValidationError("The CSV file could not be read as UTF-8 text.")
    reader = csv.reader(StringIO(decoded))
    return list(reader)


def _read_xlsx(uploaded_file):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    sheet = workbook.active
    return [
        [cell if cell is not None else "" for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]


def _read_xls(uploaded_file):
    import xlrd

    workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
    sheet = workbook.sheet_by_index(0)
    return [
        [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
        for row_index in range(sheet.nrows)
    ]


def _read_upload(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValidationError("Upload a CSV, XLSX, or XLS file.")
    if uploaded_file.size > MAX_UPLOAD_SIZE:
        raise ValidationError("The upload is too large. Please use a file smaller than 2 MB.")
    try:
        if extension == ".csv":
            return _read_csv(uploaded_file)
        if extension == ".xlsx":
            return _read_xlsx(uploaded_file)
        return _read_xls(uploaded_file)
    except Exception as exc:
        if isinstance(exc, ValidationError):
            raise
        raise ValidationError(
            "The file could not be read. Check that it is not empty or malformed."
        ) from exc


def _stringify(value):
    if value is None:
        return ""
    if hasattr(value, "date") and hasattr(value, "isoformat"):
        value = value.date()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_dict(headers, row):
    return {
        header: _stringify(row[index]) if index < len(row) else ""
        for index, header in enumerate(headers)
    }


def _choice_value(field, value):
    value = str(value).strip()
    for choice_value, choice_label in field.choices:
        if value.lower() in {str(choice_value).lower(), str(choice_label).lower()}:
            return choice_value
    return value


def _model_choice_value(field, value):
    value = str(value).strip()
    queryset = field.queryset
    for attr in ("name", "code"):
        match = queryset.filter(**{f"{attr}__iexact": value}).first()
        if match:
            return str(match.id)
    return value


def _form_data_from_row(row_values, form):
    data = {}
    labels_to_names = {_field_label(form, name): name for name in form.fields}
    for label, value in row_values.items():
        name = labels_to_names[label]
        field = form.fields[name]
        if name == "stream" and data.get("grade"):
            try:
                grade_id = UUID(str(data["grade"]))
            except ValueError:
                grade_id = None
            if grade_id is not None:
                grade = form.fields["grade"].queryset.filter(id=grade_id).first()
                if grade is not None:
                    field.queryset = field.queryset.filter(grade=grade)
        if hasattr(field, "queryset"):
            data[name] = _model_choice_value(field, value)
        elif getattr(field, "choices", None):
            data[name] = _choice_value(field, value)
        else:
            data[name] = value
    return data


def _form_errors(form):
    errors = []
    for field_name, field_errors in form.errors.as_data().items():
        label = _field_label(form, field_name) if field_name in form.fields else "Row"
        for error in field_errors:
            errors.append(BulkAdmissionError(field=label, issue=" ".join(error.messages)))
    return errors


def _field_label(form, field_name):
    field = form.fields[field_name]
    return field.label if field.label is not None else pretty_name(field_name)


def _identity(row_values):
    return (
        row_values.get("First name", "").casefold(),
        row_values.get("Middle name", "").casefold(),
        row_values.get("Last name", "").casefold(),
        row_values.get("Date of birth", ""),
    )


def parse_bulk_admission_upload(*, uploaded_file, school):
    try:
        raw_rows = _read_upload(uploaded_file)
    except ValidationError as exc:
        return BulkAdmissionResult(rows=[], file_errors=[str(message) for message in exc.messages])

    headers = bulk_admission_headers()
    if not raw_rows or not any(_stringify(cell) for cell in raw_rows[0]):
        return BulkAdmissionResult(rows=[], file_errors=["The file is empty."])
    uploaded_headers = [_stringify(cell) for cell in raw_rows[0]]
    if uploaded_headers[: len(headers)] != headers:
        return BulkAdmissionResult(
            rows=[],
            file_errors=["The file headers must match the learner admission template."],
        )

    rows = []
    seen_identities = set()
    for index, raw_row in enumerate(raw_rows[1:], start=2):
        if not any(_stringify(cell) for cell in raw_row):
            continue
        form = AdmissionForm(school=school)
        values = _row_dict(headers, raw_row)
        form = AdmissionForm(data=_form_data_from_row(values, form), school=school)
        errors = _form_errors(form) if not form.is_valid() else []
        identity = _identity(values)
        if identity in seen_identities:
            errors.append(
                BulkAdmissionError(
                    field="Learner",
                    issue="Duplicate learner inside the uploaded file.",
                )
            )
        seen_identities.add(identity)
        if Learner.objects.for_school(school).filter(
            first_name__iexact=values.get("First name", ""),
            middle_name__iexact=values.get("Middle name", ""),
            last_name__iexact=values.get("Last name", ""),
            date_of_birth=values.get("Date of birth", None) or None,
        ).exists():
            errors.append(
                BulkAdmissionError(
                    field="Learner",
                    issue="A learner with the same name and date of birth already exists.",
                )
            )
        rows.append(
            BulkAdmissionRow(
                row_number=index,
                values=values,
                cleaned_data=form.cleaned_data if form.is_valid() and not errors else {},
                errors=errors,
            )
        )
    if not rows:
        return BulkAdmissionResult(rows=[], file_errors=["The file does not contain learner rows."])
    return BulkAdmissionResult(rows=rows)


@transaction.atomic
def import_bulk_admission_rows(*, school, actor, rows):
    learners = []
    for row in rows:
        application = AdmissionApplication.objects.create(
            school=school,
            first_name=row["first_name"],
            middle_name=row.get("middle_name", ""),
            last_name=row["last_name"],
            date_of_birth=row["date_of_birth"],
            gender=row["gender"],
            desired_grade=row["grade"],
            submitted_at=row["admission_date"],
        )
        learners.append(
            admit_learner(
                school=school,
                actor=actor,
                application=application,
                academic_year=row["academic_year"],
                grade=row["grade"],
                stream=row["stream"],
                admission_date=row["admission_date"],
                learner_data={},
                guardians=[
                    {
                        "first_name": row["guardian_first_name"],
                        "last_name": row["guardian_last_name"],
                        "email": row.get("guardian_email", ""),
                        "phone_number": row["guardian_phone_number"],
                        "relationship": row["guardian_relationship"],
                        "is_primary": True,
                    }
                ],
                medical_data={
                    "blood_group": row.get("blood_group", ""),
                    "allergies": row.get("allergies", ""),
                    "conditions": row.get("conditions", ""),
                    "medication": row.get("medication", ""),
                },
            )
        )
    return learners
